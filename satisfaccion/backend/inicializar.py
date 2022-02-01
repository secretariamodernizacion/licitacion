from load import *
from portal.models import Permiso
from base.models import Institucion
from usuarios.models import User

# Permiso.objects.all().delete()

# for institucion in Institucion.objects.all():
#     user = User.objects.get_or_create(username='13961548-4')[0]
#     user.save()
#     Permiso.objects.filter(user=user, institucion=institucion).delete()
#     permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
#     permiso.tipo = 'administrador'
#     permiso.save()


user = User.objects.get(username='12813477-8')
user.set_password('1234')
user.save()
for institucion in Institucion.objects.filter(nombre__icontains='serviu'):
    permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
    permiso.tipo = 'servicio'
    permiso.save()


permiso = Permiso.objects.get_or_create(user=user, institucion=Institucion.objects.get(nombre__icontains='tesorer'))[0]
permiso.tipo = 'servicio'
permiso.save()
permiso = Permiso.objects.get_or_create(user=user, institucion=Institucion.objects.get(nombre__icontains='electric'))[0]
permiso.tipo = 'servicio'
permiso.save()


user = User.objects.get_or_create(username='8778556-4')[0]
user.set_password('1234')
user.save()
permiso = Permiso.objects.get_or_create(user=user, institucion=Institucion.objects.get(nombre__icontains='educación'))[0]
permiso.tipo = 'servicio'
permiso.save()

user = User.objects.get_or_create(username='15292863-7')[0]
user.set_password('1234')
user.save()
permiso = Permiso.objects.get_or_create(user=user, institucion=Institucion.objects.get(nombre__icontains='electri'))[0]
permiso.tipo = 'servicio'
permiso.save()


user = User.objects.get(username='9919649-1')[0]


# Instituciones
from openpyxl import load_workbook
wb = load_workbook(filename='../desarrollo/Registro Portal MESU.xlsx')
wsi = wb['Sheet1']

for row in wsi.iter_rows():
    if row[0].value == 'Servicio' or row[0].value is None:
        continue

    rut = row[2].value
    if not rut:
        continue
    rut = rut.strip().replace(".", "")
    print(rut)

    institucion = Institucion.objects.get_or_create(nombre=row[0].value)[0]
    print(institucion)
    rut = row[2].value
    if not rut:
        continue
    rut = rut.strip().replace(".", "")
    print(rut)

    user = User.objects.get_or_create(username=rut)[0]
    user.set_password('1234')
    user.save()
    Permiso.objects.filter(user=user, institucion=institucion).delete()
    permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
    permiso.tipo = 'servicio'
    permiso.save()

    user = User.objects.get_or_create(username='13961548-4')[0]
    user.save()
    Permiso.objects.filter(user=user, institucion=institucion).delete()
    permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
    permiso.tipo = 'administrador'
    permiso.save()

    user = User.objects.get_or_create(username='19475685-2')[0]
    user.save()
    Permiso.objects.filter(user=user, institucion=institucion).delete()
    permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
    permiso.tipo = 'administrador'
    permiso.save()

    user = User.objects.get_or_create(username='17088584-8')[0]
    user.save()
    permiso = Permiso.objects.get_or_create(user=user, institucion=institucion)[0]
    permiso.tipo = 'administrador'
    permiso.save()

    print(institucion)
