from load import *
from portal.models import Permiso
from base.models import Institucion, Raw
from usuarios.models import User


import xlsxwriter


from openpyxl import load_workbook
wb = load_workbook(filename='../desarrollo/Resultados2021/acnexo.xlsx')
wsi = wb[wb.sheetnames[0]]

codigo_dipres_anterior = None
raws = []
for row in wsi.iter_rows():
    if row[0].value is None:
        continue

    if row[0].value in ['Folio_unico']:
        headers = [r.value for r in row]
        continue

    cod_institucion = str(row[2].value)
    if len(cod_institucion) == 3:
        cod_institucion = '0' + cod_institucion

    # institucion = Institucion.objects.get(datos__codigo_dipres=cod_institucion)

    datos = {}
    for i, header in enumerate(headers):
        datos[header] = row[i].value

    if codigo_dipres_anterior is None or codigo_dipres_anterior != cod_institucion:
        institucion = Institucion.objects.get(codigo_dipres=cod_institucion)
        codigo_dipres_anterior = cod_institucion
        print(institucion.nombre)

    raws.append(Raw(institucion=institucion, anio=2021, datos=datos))


resp = Raw.objects.bulk_create(raws)

7 / 0
codigos = []
for raw in raws:
    if raw.codigo_cf not in codigos:
        codigos.append(raw.codigo_cf)


def limpiar(valor):
    if valor is None:
        return 0
    return int(valor)


def limpiar_pond(valor):
    if valor == '-':
        return 0
    return float(valor)


completos = []
for codigo in codigos:
    print(codigo)
    experiencia_positivo = 0
    experiencia_negativo = 0
    experiencia_total = 0
    eval_inst_positivo = 0
    eval_inst_negativo = 0
    eval_inst_total = 0
    cantidad = 0
    workbook = xlsxwriter.Workbook('../2021/{}.xlsx'.format(codigo))
    worksheet = workbook.add_worksheet()
    row = 0
    for raw in raws:
        if raw.codigo_cf != codigo:
            continue
        if row == 0:
            for col, columna in enumerate(raw.datos.keys()):
                a = worksheet.write(row, col, columna)

        for col, columna in enumerate(raw.datos.keys()):
            b = worksheet.write(row + 1, col, raw.datos.get(columna))

        row += 1
        experiencia_positivo += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PEX05")) in [6, 7] else 0
        experiencia_negativo += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PEX05")) in [1, 2, 3, 4] else 0
        experiencia_total += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PEX05")) in [1, 2, 3, 4, 5, 6, 7] else 0
        eval_inst_positivo += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PI2")) in [6, 7] else 0
        eval_inst_negativo += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PI2")) in [1, 2, 3, 4] else 0
        eval_inst_total += limpiar_pond(raw.datos.get("F_POND")) if limpiar(raw.datos.get("PI2")) in [1, 2, 3, 4, 5, 6, 7] else 0
        cantidad += 1

    experiencia_neta = round(round(experiencia_positivo / experiencia_total, 2) - round(experiencia_negativo / experiencia_total, 2), 2)
    eval_inst_neta = round(round(eval_inst_positivo / eval_inst_total, 2) - round(eval_inst_negativo / eval_inst_total, 2), 2)
    institucion = Institucion.objects.get(datos__codigo_dipres=codigo)
    data = {
        'institucion': institucion.nombre,
        'cod_institucion': codigo,
        'experiencia_positivo': experiencia_positivo,
        'experiencia_negativo': experiencia_negativo,
        'experiencia_total': experiencia_total,
        'eval_inst_positivo': eval_inst_positivo,
        'eval_inst_negativo': eval_inst_negativo,
        'eval_inst_total': eval_inst_total,
        'cantidad': cantidad,
        'experiencia_neta': experiencia_neta,
        'eval_inst_neta': eval_inst_neta,
    }
    workbook.close()

    with open('../2021/{}.json'.format(codigo), 'w') as outfile:
        json.dump(data, outfile)

    completos.append(data)


with open('../2021/acnexo.json', 'w') as outfile:
    json.dump(completos, outfile)
