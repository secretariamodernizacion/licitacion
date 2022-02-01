from load import *
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import pyreadstat

from base.models import Institucion, Raw
import json

# Raw.objects.filter().delete()


from openpyxl import Workbook
import json


wb = Workbook()
ws = wb.active
d = ws.cell(row=1, column=1, value="Institucion")
d = ws.cell(row=1, column=2, value="Año")
d = ws.cell(row=1, column=3, value="Variable")
d = ws.cell(row=1, column=4, value="Descripcion")


def importar(file, ws, row):
    anio = file.get("anio")
    nombre = file.get("nombre")
    if nombre == 'chileatiende':
        return row
    print(nombre)
    institucion = Institucion.objects.get(nombre_corto=nombre)

    file_completo = '/home/christian/repo/hacienda/satisfaccion2021/s3/metodologia/{}'.format(file.get("ubicacion"))
    try:
        df = pd.read_spss(file)
        meta = None
    except:
        df, meta = pyreadstat.read_sav(file_completo, encoding='latin1')

    # Raw.objects.filter(institucion=institucion, anio=anio).delete()

    # import ipdb
    # ipdb.set_trace()
    for label in meta.column_names:
        d = ws.cell(row=row, column=1, value=nombre)
        d = ws.cell(row=row, column=2, value=anio)
        d = ws.cell(row=row, column=3, value=label)
        d = ws.cell(row=row, column=4, value=meta.column_names_to_labels[label])
        row = row + 1

    return row
    datos = df.to_json(orient='records', lines=True).splitlines()
    raws = []
    for dato in datos:
        raws.append(Raw(institucion=institucion, anio=anio, datos=json.loads(dato)))

    # Raw.objects.bulk_create(raws)


files = [

    {"ubicacion": "2020_Bases_SPSS/2020_ESU_CHILECOMPRA.sav", "anio": 2020, "nombre": "chilecompra"},

    {"ubicacion": "2017_Bases_SPSS/2017_CHILECOMPRA_COMPRADORES_Final.sav", "anio": 2017, "nombre": "chilecompra"},
    {"ubicacion": "2017_Bases_SPSS/2017_CHILECOMPRA_PROVEEDORES_Final.sav", "anio": 2017, "nombre": "chilecompra"},

    {"ubicacion": "2015_Bases_SPSS/2015_ChileCompra_Compradores_2405.sav", "anio": 2015, "nombre": "chilecompra"},
    {"ubicacion": "2015_Bases_SPSS/2015_ChileCompra_Proveedores_2405.sav", "anio": 2015, "nombre": "chilecompra"},

    {"ubicacion": "2019_Bases_SPSS/2019 BBDD_ChileCompra Compradores_COMPLETA_COD.sav", "anio": 2019, "nombre": "chilecompra"},
    {"ubicacion": "2019_Bases_SPSS/2019 BBDD_ChileCompra Proveedores_COMPLETA_COD.sav", "anio": 2019, "nombre": "chilecompra"},

    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_ADUANAS.sav", "anio": 2020, "nombre": "aduana"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_AN.sav", "anio": 2020, "nombre": "cultura"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_CAPREDENA.sav", "anio": 2020, "nombre": "capredena"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_CHILEATIENDE_IPS.sav", "anio": 2020, "nombre": "ips"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_COMPIN.sav", "anio": 2020, "nombre": "saludpublica"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_DICREP.sav", "anio": 2020, "nombre": "dicrep"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_DIPRECA.sav", "anio": 2020, "nombre": "dipreca"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_DT.sav", "anio": 2020, "nombre": "dt"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_FONASA.sav", "anio": 2020, "nombre": "fonasa"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_FOSIS.sav", "anio": 2020, "nombre": "fosis"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_INE.sav", "anio": 2020, "nombre": "ine"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_INTERIOR.sav", "anio": 2020, "nombre": "interior"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_ISL.sav", "anio": 2020, "nombre": "isl"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_ISP.sav", "anio": 2020, "nombre": "isp"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_JUNAEB.sav", "anio": 2020, "nombre": "junaeb"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_JUNJI.sav", "anio": 2020, "nombre": "junji"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_MINVU.sav", "anio": 2020, "nombre": "vivienda"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SEC.sav", "anio": 2020, "nombre": "sec"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SENAMA.sav", "anio": 2020, "nombre": "senama"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SENCE.sav", "anio": 2020, "nombre": "sence"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SERNAC.sav", "anio": 2020, "nombre": "sernac"},
    # # {"ubicacion":"2020_Bases_SPSS/2020_ESU_SERNAC_old.sav", "anio":2020, "nombre": ""},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SERNAPESCA.sav", "anio": 2020, "nombre": "sernapesca"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SII.sav", "anio": 2020, "nombre": "sii"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SRCEI.sav", "anio": 2020, "nombre": "registrocivil"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SUBS_ECON.sav", "anio": 2020, "nombre": "economia"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SUBS_SERV_SOCIALES.sav", "anio": 2020, "nombre": "serviciossociales"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SUPER_PENSIONES.sav", "anio": 2020, "nombre": "superpensiones"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SUPER_SALUD.sav", "anio": 2020, "nombre": "supersalud"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_SUSESO.sav", "anio": 2020, "nombre": "suseso"},
    # {"ubicacion": "2020_Bases_SPSS/2020_ESU_TESORERIA.sav", "anio": 2020, "nombre": "tesoreria"},

    # {"ubicacion": "2019_Bases_SPSS/2019 Base DT.sav", "anio": 2019, "nombre": "dt"},
    # {"ubicacion": "2019_Bases_SPSS/2019 Base ISL.sav", "anio": 2019, "nombre": "isl"},
    # {"ubicacion": "2019_Bases_SPSS/2019 Base MINVU.sav", "anio": 2019, "nombre": "vivienda"},
    # {"ubicacion": "2019_Bases_SPSS/2019 Base SuperSalud.sav", "anio": 2019, "nombre": "supersalud"},
    # {"ubicacion": "2019_Bases_SPSS/2019 Base SUSESO.sav", "anio": 2019, "nombre": "suseso"},
    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_ChileAtiende_COMPLETA_COD.sav", "anio": 2019, "nombre": "chileatiende"},

    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_FOSIS_COMPLETA_COD.sav", "anio": 2019, "nombre": "fosis"},
    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_INE_COMPLETA_COD.sav", "anio": 2019, "nombre": "ine"},
    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_Registro Civil_COMPLETA_COD.sav", "anio": 2019, "nombre": "registrocivil"},
    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_SERNAC_COMPLETA_COD.sav", "anio": 2019, "nombre": "sernac"},
    # {"ubicacion": "2019_Bases_SPSS/2019 BBDD_Subse.Eco_COMPLETA.sav", "anio": 2019, "nombre": "economia"},
    # {"ubicacion": "2019_Bases_SPSS/2019_BBDD_TGR_COMPLETA.sav", "anio": 2019, "nombre": "tesoreria"},
    # {"ubicacion": "2019_Bases_SPSS/2019 COMPIN.sav", "anio": 2019, "nombre": "saludpublica"},
    # {"ubicacion": "2019_Bases_SPSS/2019 FONASA.sav", "anio": 2019, "nombre": "fonasa"},
    # {"ubicacion": "2019_Bases_SPSS/2019 IPS.sav", "anio": 2019, "nombre": "ips"},
    # {"ubicacion": "2019_Bases_SPSS/2019 SuperPensiones.sav", "anio": 2019, "nombre": "superpensiones"},

    # {"ubicacion": "2018_Bases_SPSS/2018 01 Base SUSESO.sav", "anio": 2018, "nombre": "suseso"},
    # {"ubicacion": "2018_Bases_SPSS/2018 02 Base SII.sav", "anio": 2018, "nombre": "sii"},
    # {"ubicacion": "2018_Bases_SPSS/2018 03 Base COMPIN.sav", "anio": 2018, "nombre": "saludpublica"},
    # {"ubicacion": "2018_Bases_SPSS/2018 04 Base AN.sav", "anio": 2018, "nombre": "cultura"},
    # {"ubicacion": "2018_Bases_SPSS/2018 05 Base IPS.sav", "anio": 2018, "nombre": "ips"},
    # # {"ubicacion": "2018_Bases_SPSS/2018 06 Base SRCeI Instituciones.sav", "anio": 2018, "nombre": "registrocivil"},
    # {"ubicacion": "2018_Bases_SPSS/2018 06 Base SRCeI Usuario.sav", "anio": 2018, "nombre": "registrocivil"},
    # {"ubicacion": "2018_Bases_SPSS/2018 07 Base FONASA.sav", "anio": 2018, "nombre": "fonasa"},
    # {"ubicacion": "2018_Bases_SPSS/2018 08 Base DT.sav", "anio": 2018, "nombre": "dt"},
    # {"ubicacion": "2018_Bases_SPSS/2018 09 Base TGR.sav", "anio": 2018, "nombre": "tesoreria"},

    # {"ubicacion": "2017_Bases_SPSS/2017_BASE DATOS LINEA BASE CHILEATIENDE.sav", "anio": 2017, "nombre": "chileatiende"},
    # {"ubicacion": "2017_Bases_SPSS/2017_CHILEATIENDE_FICHA_Final.sav", "anio": 2017, "nombre": "chileatiende"},
    # {"ubicacion": "2017_Bases_SPSS/2017_CHILEATIENDE_HOME_Final.sav", "anio": 2017, "nombre": "chileatiende"},
    # {"ubicacion": "2017_Bases_SPSS/2017_INE_Final.sav", "anio": 2017, "nombre": "ine"},
    # {"ubicacion": "2017_Bases_SPSS/2017_REGISTRO CIVIL_Final.sav", "anio": 2017, "nombre": "registrocivil"},
    # {"ubicacion": "2017_Bases_SPSS/2017_SERNAC_Reclamos y Mediaciones 07-12.sav", "anio": 2017, "nombre": "sernac"},

    # {"ubicacion": "2016_Bases_SPSS/201612_170502_Informe 4 Base DT_consolidado.sav", "anio": 2016, "nombre": "dt"},
    # {"ubicacion": "2016_Bases_SPSS/201612_170502_Informe 4 Base SUSESO_consolidado.sav", "anio": 2016, "nombre": "suseso"},
    # {"ubicacion": "2016_Bases_SPSS/201612_170519_Informe 4 Base SII_consolidado.sav", "anio": 2016, "nombre": "sii"},

    # {"ubicacion": "2015_Bases_SPSS/2015_BaseDatos_ChileAtiende Oficinas_2405.sav", "anio": 2015, "nombre": "chileatiende"},
    # {"ubicacion": "2015_Bases_SPSS/2015_BaseDatos_ChileAtiende_Web_2405.sav", "anio": 2015, "nombre": "chileatiende"},
    # {"ubicacion": "2015_Bases_SPSS/2015_BaseDatos_INE_2405.sav", "anio": 2015, "nombre": "ine"},
    # {"ubicacion": "2015_Bases_SPSS/2015_BaseDatos_SERNAC_0106_v2.sav", "anio": 2015, "nombre": "sernac"},
]

row = 2
for file in files:
    row = importar(file, ws, row)


wb.save('columns.xlsx')
