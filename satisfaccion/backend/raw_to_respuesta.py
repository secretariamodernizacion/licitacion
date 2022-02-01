import sys
import os
import importlib

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
os.environ['DJANGO_SETTINGS_MODULE'] = 'conf.settings'
application = get_wsgi_application()

from base.models import *
Institucion.objects.all().delete()

# Instituciones
from openpyxl import load_workbook
wb = load_workbook(filename='Instituciones.xlsx')
wsi = wb['Sheet1']

for row in wsi.iter_rows():
    if row[0].value == 'ID':
        continue
    institucion = Institucion.objects.get_or_create(pk=row[0].value)[0]
    print(row[2].value.strip())
    institucion.datos["agrupacion"] = (row[7].value or "").strip()
    institucion.datos["codigo"] = (row[2].value or "").strip()
    institucion.datos["order"] = int(row[9].value)
    institucion.datos["nombre_en_data"] = (row[8].value or "").strip()
    institucion.nombre = (row[1].value or "").strip()
    institucion.save()
    print(institucion.id)

resp = []
for grupo in Institucion.objects.filter().distinct('datos__agrupacion').order_by('datos__agrupacion').values('datos__agrupacion'):
    print(grupo.get("datos__agrupacion"))
    instituciones = []
    instituciones_q = Institucion.objects.filter(datos__agrupacion=grupo.get("datos__agrupacion")).order_by('datos__orden')
    print(len(instituciones_q))
    for institucion in instituciones_q:
        instituciones.append({'nombre': institucion.nombre,
                              'codigo': institucion.datos.get("codigo"),
                              'id': institucion.id})
    resp.append({'nombre': grupo.get("datos__agrupacion", 'Sin nombre'), 'instituciones': instituciones})

import json
with open("instituciones_por_grupo.json", "w") as write_file:
    json.dump(resp, write_file, indent=2)


from base.models import Respuesta, Raw, Institucion

# for k in range(50):
#     try:
#         # institucion = Institucion.objects.get_or_create(nombre='ChileaAtiende - Presencial')[0]
#         # institucion.datos["codigo"] = institucion.nombre
#         # institucion.save()

#         # institucion = Institucion.objects.get_or_create(nombre='ChileaAtiende - Digital')[0]
#         # institucion.datos["codigo"] = institucion.nombre
#         # institucion.save()

#         institucion = Institucion.objects.get_or_create(nombre='ChileaAtiende')[0]
#         institucion.datos["codigo"] = institucion.nombre
#         institucion.save()
#     except:
#         pass


def cambiar_nombre(codigo_cf_nombre):
    codigo_cf_nombre = codigo_cf_nombre.replace("SERVIU-", "SERVIU ")
    if codigo_cf_nombre == 'SERVIU Aysén del Gral. Carlos Ibáñez del Campo':
        codigo_cf_nombre = 'SERVIU Aysén'
    if codigo_cf_nombre == 'SERIVU Arica y Parinacota':
        codigo_cf_nombre = 'SERVIU Arica y Parinacota'

    if codigo_cf_nombre.lower().find("atiend") >= 0:
        codigo_cf_nombre = 'ChileAtiende'

    # if codigo_cf_nombre == 'CHILECOMPRA-COMPRADORES':
    #     codigo_cf_nombre = 'Chilecompra-Compradores'
    # if codigo_cf_nombre == 'CHILEATIENDE_PRESENCIAL':
    #     codigo_cf_nombre = 'ChileAtiende'
    if codigo_cf_nombre == 'SRCeI':
        codigo_cf_nombre = 'SRC'
    if codigo_cf_nombre == 'SUBSEECONOMIA':
        codigo_cf_nombre = 'SubseEco'
    if codigo_cf_nombre == 'SUPER-PENSIONES':
        codigo_cf_nombre = 'SuperPensiones'
    if codigo_cf_nombre == 'TGR':
        codigo_cf_nombre = 'Tesoreria'

    if codigo_cf_nombre == 'SERVIU TARAPACA':
        codigo_cf_nombre = 'SERVIU TARAPACá'
    if codigo_cf_nombre == 'SERVIU VALPARAISO':
        codigo_cf_nombre = 'SERVIU VALPARAíSO'

    if codigo_cf_nombre == 'SERVIU OHIGGINS':
        codigo_cf_nombre = "SERVIU O'Higgins"
    if codigo_cf_nombre == 'SERVIU BIOBIO':
        codigo_cf_nombre = 'Serviu Biobío'
    if codigo_cf_nombre == 'SERVIU ARAUCANIA':
        codigo_cf_nombre = 'SERVIU ARAUCANíA'
    if codigo_cf_nombre == 'SERVIU LOSLAGOS':
        codigo_cf_nombre = 'SERVIU Los lagos'
    if codigo_cf_nombre == 'SERVIU AYSEN':
        codigo_cf_nombre = 'SERVIU Aysén'
    if codigo_cf_nombre == 'SERVIU METROPOLITANO':
        codigo_cf_nombre = 'SERVIU METROPOLITANa'
    if codigo_cf_nombre == 'SERVIU LOSRIOS':
        codigo_cf_nombre = 'SERVIU Los ríos'
    if codigo_cf_nombre == 'SERVIU ARICA-PARINACOTA':
        codigo_cf_nombre = 'SERVIU Arica y parinacota'

    if codigo_cf_nombre == 'SERVIU Nuble':
        codigo_cf_nombre = 'SERVIU Ñuble'

    return codigo_cf_nombre


def nivel_educacional(nivel):
    try:
        nivel = nivel.strip()
    except:
        pass
    if nivel == '3.- Superior Completa':
        return 'superior'
    if nivel == '1.- Escolar Incompleta':
        return 'escolar_incompleta'
    if nivel == '2.- Escolar Completa':
        return 'escolar_completa'


def sexo_transformar(sexo):
    if sexo == 'Femenino':
        return 'mujer'
    if sexo == 'Masculino':
        return 'hombre'


def toInteger(val):
    if val is None:
        return None

    try:
        val = val.strip()
    except:
        pass

    if val == '1 (pésimo)':
        return 1

    if val == '7 (excelente)':
        return 7

    if val == '':
        return None

    try:
        return int(val)
    except:
        print("Problema en {}".format(val))
        return None


def pasar_a_nombre(valor):
    try:
        valor = valor.strip()
    except:
        pass
    try:
        return valor.split(".- ")[-1].replace(" ", "").lower()
    except:
        return valor


Respuesta.objects.all().delete()

respuestas = []
prev_institucion = ''

for codigo_cf_item in Raw.objects.all().distinct('codigo_cf').values('codigo_cf'):
    codigo_cf = codigo_cf_item["codigo_cf"]
    codigo_cf_nombre = cambiar_nombre(codigo_cf)
    print(codigo_cf)
    institucion = Institucion.objects.filter(datos__codigo__iexact=codigo_cf_nombre.strip()).first()
    for row in Raw.objects.filter(codigo_cf=codigo_cf):
        respuesta = Respuesta(
            institucion=institucion,
            anio=row.anio,
            experiencia=row.datos.get("PEX_eval"),
            eval_inst=row.datos.get("PI_eval_inst"),
            f_pond=row.datos.get("F_POND"),
            edad=row.datos.get("U_edad_agrupada").strip(),
            sexo=sexo_transformar(row.datos.get("U_sexo").strip()),
            educacion=nivel_educacional(row.datos.get("U_educacion_agr")),
            habilitado=pasar_a_nombre(row.datos.get("PHAB_0")),
            tipo_tramite=row.datos.get("T.Tramite"),

            pi_imagen_confianza=toInteger(row.datos.get("PI_imagen_confianza")),
            pi_imagen_transp=toInteger(row.datos.get("PI_imagen_transp")),
            pi_imagen_preocupa=toInteger(row.datos.get("PI_imagen_preocupa")),
            pi_imagen_actual=toInteger(row.datos.get("PI_imagen_actual")),
            pi_imagen_funcion=toInteger(row.datos.get("PI_imagen_funcion")),
            pi_imagen_satisface=toInteger(row.datos.get("PI_imagen_satisface")),
            pev_tram_facil=toInteger(row.datos.get("PEV_tram_facil")),
            pev_tram_tiempo=toInteger(row.datos.get("PEV_tram_tiempo")),
            pev_tram_claridad_pasos=toInteger(row.datos.get("PEV_tram_claridad_pasos")),
            pev_tram_info_compr=toInteger(row.datos.get("PEV_tram_info_compr")),
            pev_tram_info_util=toInteger(row.datos.get("PEV_tram_info_util")),
            pev_tram_resp_util=toInteger(row.datos.get("PEV_tram_resp_util")),
            pev_tram_resp_compl=toInteger(row.datos.get("PEV_tram_resp_compl")),
            pev_tram_resp_clara=toInteger(row.datos.get("PEV_tram_resp_clara")),
            pev_tram_acogido=toInteger(row.datos.get("PEV_tram_acogido")),

            pec_eval_telefono=toInteger(row.datos.get("PEC_eval_telefono")),
            pec_eval_web=toInteger(row.datos.get("PEC_eval_web")),
            pec_eval_presencial=toInteger(row.datos.get("PEC_eval_presencial")),

            pec_canal_telefono=1 if row.datos.get("PEC_canal_telefono") and str(row.datos.get("PEC_canal_telefono")) != 0 else 0,
            pec_canal_web=1 if row.datos.get("PEC_canal_web") and str(row.datos.get("PEC_canal_web")) != 0 else 0,
            pec_canal_presencial=1 if row.datos.get("PEC_canal_presencial") and str(row.datos.get("PEC_canal_presencial")) != 0 else 0,
        )
        respuesta.save()
        # respuestas.append(respuesta)

    # r = Respuesta.objects.bulk_create(respuestas)


def get_pec(valor):
    try:
        valor = int(float(valor))
    except:
        pass
    if valor.lower() in ['si', 'sí', 1]:
        return 1
    if valor.lower() in ['no', 0]:
        return 0


for institucion in Institucion.objects.all():
    anios = [i.get("anio") for i in institucion.respuesta_set.all().distinct('anio').values('anio')]
    institucion.datos["anios"] = anios
    if len(anios) == 0:
        institucion.datos["anios_string"] = None
    if len(anios) == 1:
        institucion.datos["anios_string"] = anios[0]

    if len(anios) > 1:
        institucion.datos["anios_string"] = ", ".join([str(a) for a in anios][0:-1]) + " y " + str(anios[-1])

    institucion.save()

for institucion in Institucion.objects.all():
    institucion.resumen = institucion.calcular_resumen()
    institucion.save()


institucion = Institucion.objects.get(datos__codigo__iexact='Chileatiende')
institucion.resumen = institucion.calcular_resumen()
institucion.save()

Respuesta.objects.filter(educacion='escolar_incompleta').update(educacion='Escolar incompleta')
Respuesta.objects.filter(educacion='escolar_completa').update(educacion='Escolar completa')
Respuesta.objects.filter(educacion='superior').update(educacion='Superior')

Respuesta.objects.filter(habilitado='habilitado').update(habilitado='Habilitado')
Respuesta.objects.filter(habilitado='medianamentehabilitado').update(habilitado='Medianamente habilitado')
Respuesta.objects.filter(habilitado='nohabilitado').update(habilitado='No habilitado')


inst = Institucion.objects.get(nombre='Dirección del Trabajo')
inst.datos["agrupacion"] = 'Fiscalización'
inst.save()

inst = Institucion.objects.get(nombre='Subsecretaría de Salud Pública (COMPIN)')
inst.datos["agrupacion"] = 'Salud/Previsión'
inst.save()


inst = Institucion.objects.get(nombre='Archivo Nacional')
inst.datos["agrupacion"] = 'Ciudadanía /T. Generales'
inst.save()
